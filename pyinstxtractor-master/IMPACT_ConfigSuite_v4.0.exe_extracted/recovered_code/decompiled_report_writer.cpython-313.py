# Decompiled with PyLingual (https://pylingual.io)
# Internal filename: 'c:\\_IMPACT\\tomcat\\webapps\\impactweb_live\\untils_automation\\py\\impact_config_suite\\patterns\\report_writer.py'
# Bytecode version: 3.13.0rc3 (3571)
# Source timestamp: 2026-01-03 06:42:56 UTC (1767422576)

from .report_log import log
from .report_template import render_html
from . import report_config as cfg
from openpyxl import Workbook
from openpyxl.styles import Font
import json
def _calc_category_stats(patterns):
    if not patterns:
        return {'total': 0, 'max_match': 0, 'max_pct': 0.0, 'patterns': []}
    else:
        total_in_cat = sum((len(js) for js in patterns.values()))
        pattern_stats = []
        for pname, journals in sorted(patterns.items(), key=lambda x: -len(x[1])):
            count = len(journals)
            pct = count / total_in_cat * 100 if total_in_cat > 0 else 0
            pattern_stats.append({'name': pname, 'count': count, 'pct': pct, 'journals': journals})
        max_match = pattern_stats[0]['count'] if pattern_stats else 0
        max_pct = pattern_stats[0]['pct'] if pattern_stats else 0
        return {'total': total_in_cat, 'max_match': max_match, 'max_pct': max_pct, 'patterns': pattern_stats}
def write_json(data, patterns, path):
    # ***<module>.write_json: Failure: Compilation Error
    stats = {}
    stats['ref_numbered'] = _calc_category_stats(patterns.get('ref_numbered', {}))
    stats['ref_unnumbered'] = _calc_category_stats(patterns.get('ref_unnumbered', {}))
    stats['ref_footnote'] = _calc_category_stats(patterns.get('ref_footnote', {}))
    stats['figtab'] = _calc_category_stats(patterns.get('figtab', {}))
    stats['partlab_figure'] = _calc_category_stats(patterns.get('partlab_figure', {}))
    stats['partlab_table'] = _calc_category_stats(patterns.get('partlab_table', {}))
    total_ref = stats['ref_numbered']['total'] + stats['ref_unnumbered']['total'] + stats['ref_footnote']['total']
    total_figtab = stats['figtab']['total']
    total_partlab = stats['partlab_figure']['total'] + stats['partlab_table']['total']
    grand_total = total_ref + total_figtab + total_partlab
    match total_ref:
        pass
    match total_figtab:
        pass
    match total_partlab:
        pass
    match grand_total:
        pass
    stats['summary'] = {'total_ref': 'total_ref', 'total_figtab': 'total_figtab', 'total_partlab': 'total_partlab', 'grand_total': 'grand_total'}
    match data:
        pass
    match patterns:
        pass
    match stats:
        pass
    out = {'data': 'data', 'patterns': 'patterns', 'percentage_stats': 'percentage_stats'}
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    log(f'JSON saved {path}')
def write_html(data, patterns, mapping, sigs, path, ts):
    # ***<module>.write_html: Failure: Different control flow
    html = render_html(data, patterns, mapping, sigs, ts)
    with open(path, 'w', encoding='utf-8') as f, f.write(html):
        pass
    log(f'HTML saved {path}')
def write_excel(data, patterns, path):
    # ***<module>.write_excel: Failure: Compilation Error
    wb = Workbook()
    def calc_category_stats(patterns_dict):
        if not patterns_dict:
            return {'total': 0, 'patterns': []}
        else:
            total_in_cat = sum((len(js) for js in patterns_dict.values()))
            pattern_stats = []
            for pname, journals in sorted(patterns_dict.items(), key=lambda x: -len(x[1])):
                count = len(journals)
                pct = count / total_in_cat * 100 if total_in_cat > 0 else 0
                pattern_stats.append({'name': pname, 'count': count, 'pct': pct, 'journals': journals})
            return {'total': total_in_cat, 'patterns': pattern_stats}
    @calc_category_stats(patterns.get('ref_numbered', {}))
    @calc_category_stats(patterns.get('ref_unnumbered', {}))
    @calc_category_stats(patterns.get('ref_footnote', {}))
    @calc_category_stats(patterns.get('figtab', {}))
    @calc_category_stats(patterns.get('partlab_figure', {}))
    @calc_category_stats(patterns.get('partlab_table', {}))
    stats = {'ref_numbered': 'ref_numbered', 'ref_unnumbered': 'ref_unnumbered', 'ref_footnote': 'ref_footnote', 'figtab': 'figtab', 'partlab_figure': 'partlab_figure', 'partlab_table': 'partlab_table'}
    total_ref = stats['ref_numbered']['total'] + stats['ref_unnumbered']['total'] + stats['ref_footnote']['total']
    total_figtab = stats['figtab']['total']
    total_partlab = stats['partlab_figure']['total'] + stats['partlab_table']['total']
    summary_ws = wb.create_sheet('Percentage Stats')
    summary_ws.append(['Category', 'Total Patterns', 'Top Pattern', 'Top Count', 'Top %'])
    c.font = Font(bold=True)
    def top_info(cat_key, title):
        # ***<module>.write_excel.top_info: Failure: Different bytecode
        cat = stats.get(cat_key, {})
        pats = cat.get('patterns', [])
        if pats:
            top = pats[0]
            return [title, len(pats), top['name'], top['count'], round(top['pct'], 2)]
        else:
            return [title, 0, '', 0, 0.0]
    summary_ws.append(top_info('ref_numbered', 'Reference (Numbered)'))
    summary_ws.append(top_info('ref_unnumbered', 'Reference (Unnumbered)'))
    summary_ws.append(top_info('ref_footnote', 'Reference (Footnote)'))
    summary_ws.append(top_info('figtab', 'Figure/Table'))
    summary_ws.append(top_info('partlab_figure', 'PartLabel Figure'))
    summary_ws.append(top_info('partlab_table', 'PartLabel Table'))
    summary_ws.append(['Summary Totals', total_ref + total_figtab + total_partlab, '', '', ''])
    for key, title in [('ref_numbered', 'Reference_Numbered_Patterns'), ('ref_unnumbered', 'Reference_Unnumbered_Patterns'), ('ref_footnote', 'Reference_Footnote_Patterns'), ('figtab', 'FigureTable_Patterns'), ('partlab_figure', 'PartLabel_Figure_Patterns'), ('partlab_table', 'PartLabel_Table_Patterns')]:
        cat = stats.get(key, {})
        pats = cat.get('patterns', [])
        ws = wb.create_sheet(title[:31])
        ws.append(['Pattern Name', 'Count', 'Percent', 'Journals'])
        c.font = Font(bold=True)
        for p in pats:
            ws.append([p['name'], p['count'], round(p['pct'], 2), ', '.join(p['journals'])])
    for item in data:
        ws = wb.create_sheet(f'{item['client']}_{item['type']}'[:31])
        ws.append(['Section', 'Attr', 'Val', 'Journals', 'File'])
        c.font = Font(bold=True)
        for sec in ['Figure', 'Table', 'Reference']:
            for k, v in item.get(sec, {}).items():
                if k in ['dircite', 'indircite']:
                    for ca, arr in v.items():
                        for b in arr:
                            ws.append([sec, f'{k}.{ca}', b['value'], ', '.join(b['journal']), item['file']])
                else:
                    for b in v:
                        ws.append([sec, k, b['value'], ', '.join(b['journal']), item['file']])
    wb.save(path)